from itertools import count
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout, get_user_model
from django.contrib.auth.decorators import login_required, user_passes_test
from django.views import View
from .forms import CategoriaForm
from .forms import TipoMaterialForm
from .forms import MarcaForm
from .forms import MaterialForm
from .models import Material, Carrito, ItemCarrito, Solicitud, ItemSolicitud, CategoriaDj, Marca, TipoMaterial
from .forms import LoginForm, CrearUsuarioForm
from django.contrib import messages
from .forms import GestionarSolicitudForm
from .forms import SolicitudForm
from .forms import EditarPerfilForm
from .forms import DevolverItemForm
from django.db.models import Q
from django.contrib.admin.views.decorators import staff_member_required
from django.utils.timezone import now
from django.template.loader import get_template
from xhtml2pdf import pisa
from django.http import HttpResponse
from io import BytesIO
from django.urls import reverse
from django.http import JsonResponse
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash
from django.contrib import messages
from time import timezone
from datetime import datetime, date, timedelta
from django.core.mail import EmailMultiAlternatives
from appSIGEM.templatetags.utils import enviar_correo_admin
from appSIGEM.templatetags.utils import enviar_correo_usuario_respuesta
from django.core.mail import send_mail
from django.conf import settings
from django.contrib.auth import get_user_model
from django.contrib.sites.shortcuts import get_current_site
from django.template.loader import render_to_string
from django.utils.http import urlsafe_base64_encode
from django.utils.encoding import force_bytes
from django.contrib.auth.tokens import default_token_generator
from django.utils.http import urlsafe_base64_decode
from django.core.mail import EmailMessage
from django.http import HttpResponseRedirect
from django.core.paginator import Paginator
from django.contrib.messages import get_messages
import requests
from .models import SlideCarrusel
from .forms import SlideCarruselForm


User = get_user_model()  # Obtiene el modelo de usuario actual de Django

# Función que verifica si el usuario es un superusuario
def is_admin(user):
    return user.is_superuser


class LoginView(View):
    def get(self, request):
        # Si el usuario ya está autenticado, redirige a la página correspondiente
        if request.user.is_authenticated:
            if request.user.is_superuser:
                return redirect('admin_panel')  # Redirigir al panel de administrador si es superusuario
            return redirect('index')  # Redirigir al índice si es usuario normal
        
        form = LoginForm()
        return render(request, 'paginas/login/login.html', {'form': form})

    def post(self, request):
        form = LoginForm(request, data=request.POST)
        
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']
            user = authenticate(request, username=username, password=password)
            
            if user is not None:
                login(request, user)
                if user.is_superuser:
                    return redirect('admin_panel')
                return redirect('index')
            else:
                # Usuario ingresado pero contraseña incorrecta o usuario inexistente
                user_exists = User.objects.filter(username=username).exists()
                if user_exists:
                    error_message = "La contraseña no es correcta."
                else:
                    error_message = "El usuario no existe."
                return render(request, 'paginas/login/login.html', {
                    'form': form,
                    'error': error_message
                })
        else:
            # Formulario no válido: puede deberse a campos vacíos u otros errores
            return render(request, 'paginas/login/login.html', {
                'form': form,
                'error': "Usuario y/o contraseña no válidos."
            })

@login_required
def index(request):
    api_key = '251ce61d7fd44f4395658df3e3606ff5'
    url = f'https://newsapi.org/v2/everything?q=tecnología&language=es&sortBy=publishedAt&pageSize=6&apiKey={api_key}'

    response = requests.get(url)
    noticias = []

    if response.status_code == 200:
        data = response.json()
        noticias = data.get('articles', [])[:3]
    else:
        print("Error al obtener noticias:", response.status_code)

    slides = SlideCarrusel.objects.all()

    return render(request, 'paginas/inicio/index.html', {
        'noticias': noticias,
        'slides': slides,
    })
# Vista de logout
@login_required
def logout_view(request):
    logout(request)
    return redirect('login')

# Vista solo para superusuarios: crear nuevo usuario

def crear_usuario(request):
    if request.method == 'POST':
        form = CrearUsuarioForm(request.POST, request.FILES)
        if form.is_valid():
            user = form.save(commit=False)
            user.is_active = False  # Usuario inactivo hasta verificar correo
            user.save()

            current_site = get_current_site(request)
            mail_subject = 'Activa tu cuenta en SIGEM'
            uid = urlsafe_base64_encode(force_bytes(user.pk))
            token = default_token_generator.make_token(user)
            activation_link = f"http://{current_site.domain}/activar/{uid}/{token}/"
            
            message = render_to_string('emails/activar_cuenta.html', {
                'user': user,
                'activation_link': activation_link,
            })
            email = EmailMessage(mail_subject, message, to=[user.email])
            email.content_subtype = 'html'  # Enviar como HTML
            email.send()

            return render(request, 'paginas/registro_pendiente_verificacion.html', {
                'email': user.email
            })
    else:
        form = CrearUsuarioForm()

    return render(request, 'paginas/login/crear_usuario.html', {'form': form})


def enviar_correo_confirmacion(user):
    asunto = 'Bienvenido a SIGEM'
    mensaje = f'''
    Hola {user.first_name},

    Tu cuenta en SIGEM ha sido creada correctamente.

    Usuario: {user.username}
    Correo: {user.email}

    Puedes iniciar sesión aquí: http://127.0.0.1:8000/login/

    Saludos,
    Equipo SIGEM
    '''
    send_mail(asunto, mensaje, settings.DEFAULT_FROM_EMAIL, [user.email])
    
    
def enviar_correo_verificacion(request, user):
    current_site = get_current_site(request)
    uid = urlsafe_base64_encode(force_bytes(user.pk))
    token = default_token_generator.make_token(user)
    activation_link = f"http://{current_site.domain}/activar-cuenta/{uid}/{token}/"

    subject = "Activa tu cuenta en SIGEM"
    message = render_to_string('emails/activar_cuenta.html', {
        'user': user,
        'activation_link': activation_link,
    })

    send_mail(subject, message, settings.DEFAULT_FROM_EMAIL, [user.email])
    
    
def activar_cuenta(request, uidb64, token):
    try:
        uid = urlsafe_base64_decode(uidb64).decode()
        user = get_object_or_404(User, pk=uid)
    except (TypeError, ValueError, OverflowError, User.DoesNotExist):
        user = None

    if user and default_token_generator.check_token(user, token):
        user.is_active = True
        user.save()
        messages.success(request, 'Cuenta activada correctamente. Ya puedes iniciar sesión.')
        return redirect('login')  # Asegúrate de que exista la URL 'login'
    else:
        return render(request, 'registro/activacion_invalida.html')
    
def enviar_email_html(destinatario, context):
    subject = 'Restablecer Contraseña - SIGEM'
    from_email = 'noreply@sigem.com'
    to = [destinatario]

    html_content = render_to_string('paginas/auth/password_reset_email.html', context)
    text_content = f"""
    Hola {context['user'].get_full_name() or context['user'].username},
    Has solicitado restablecer tu contraseña en SIGEM.
    Puedes hacerlo ingresando al siguiente enlace:
    {context['reset_link']}
    """

    msg = EmailMultiAlternatives(subject, text_content.strip(), from_email, to)
    msg.attach_alternative(html_content, "text/html")
    msg.send()
    
def solicitud_reset_password(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        try:
            user = User.objects.get(email=email)
            current_site = get_current_site(request)
            domain = current_site.domain
            protocol = 'https' if request.is_secure() else 'http'

            uid = urlsafe_base64_encode(force_bytes(user.pk))
            token = default_token_generator.make_token(user)

            reset_link = f"{protocol}://{domain}/reset/{uid}/{token}/"

            context = {
                'user': user,
                'reset_link': reset_link,
                'domain': f"{protocol}://{domain}"
            }

            enviar_email_html(email, context)
            messages.success(request, 'Correo de restablecimiento enviado. Revisa tu bandeja de entrada.')
            return redirect('login')

        except User.DoesNotExist:
            messages.error(request, 'No hay una cuenta registrada con ese correo.')
    
    return render(request, 'paginas/auth/password_reset_form.html')

# Vista para administrador
@login_required
@user_passes_test(is_admin)
def admin_panel(request):
    return render(request, 'paginas/inicio/admin_panel.html')

@login_required
def perfil_usuario(request):
    # Asumiendo que tienes un modelo de perfil asociado al usuario
    # Se puede acceder directamente a la imagen de perfil desde el modelo User si es que la has añadido allí
    user = request.user  # Obtén el usuario actualmente autenticado
    return render(request, 'paginas/inicio/perfil.html', {'user': user})
# Agregar categoría
def agregar_categoria(request):
    if request.method == 'POST':
        form = CategoriaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Categoría agregada exitosamente.')
            return redirect('agregar_categoria')  # Puedes redirigir a otra vista si prefieres
    else:
        form = CategoriaForm()
    return render(request, 'paginas/agregar_cosas/agregar_categoria.html', {'form': form})

#


# Agregar tipo de material
def agregar_tipo_material(request):
    if request.method == 'POST':
        form = TipoMaterialForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Tipo de material agregado exitosamente.')
            return redirect('agregar_tipo_material')
    else:
        form = TipoMaterialForm()
    return render(request, 'paginas/agregar_cosas/agregar_tipo_material.html', {'form': form})


# Agregar marca
def agregar_marca(request):
    if request.method == 'POST':
        form = MarcaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Marca agregada exitosamente.')
            return redirect('agregar_marca')
    else:
        form = MarcaForm()
    return render(request, 'paginas/agregar_cosas/agregar_marca.html', {'form': form})


# Agregar material
def agregar_material(request):
    if request.method == 'POST':
        form = MaterialForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Material agregado exitosamente.')
            return redirect('agregar_material')
    else:
        form = MaterialForm()
    return render(request, 'paginas/agregar_cosas/agregar_material.html', {'form': form})

#listar materiales

from django.contrib.auth.decorators import login_required
from .models import Material, CategoriaDj



from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.shortcuts import render
from .models import Material, CategoriaDj, Carrito, ItemSolicitud

@login_required
def listar_materiales(request):
    hoy = date.today()

    # Si no hay parámetros GET, limpiar filtros de sesión
    if not request.GET:
        for key in ['q_filtro', 'marca_filtro', 'tipo_filtro', 'categoria_filtro', 'fecha_inicio_filtro', 'fecha_fin_filtro']:
            request.session.pop(key, None)

    # Manejo de filtros con sesión
    q = request.GET.get('q')
    if q is not None:
        q = q.strip()
        request.session['q_filtro'] = q
    else:
        q = request.session.get('q_filtro', '')

    marca = request.GET.get('marca')
    if marca is not None:
        marca = marca.strip()
        request.session['marca_filtro'] = marca
    else:
        marca = request.session.get('marca_filtro', '')

    tipo = request.GET.get('tipo')
    if tipo is not None:
        tipo = tipo.strip()
        request.session['tipo_filtro'] = tipo
    else:
        tipo = request.session.get('tipo_filtro', '')

    categoria = request.GET.get('categoria')
    if categoria is not None:
        categoria = categoria.strip()
        request.session['categoria_filtro'] = categoria
    else:
        categoria = request.session.get('categoria_filtro', '')

    fecha_inicio_str = request.GET.get('fecha_inicio')
    if fecha_inicio_str is not None:
        fecha_inicio_str = fecha_inicio_str.strip()
        request.session['fecha_inicio_filtro'] = fecha_inicio_str
    else:
        fecha_inicio_str = request.session.get('fecha_inicio_filtro', '')

    fecha_fin_str = request.GET.get('fecha_fin')
    if fecha_fin_str is not None:
        fecha_fin_str = fecha_fin_str.strip()
        request.session['fecha_fin_filtro'] = fecha_fin_str
    else:
        fecha_fin_str = request.session.get('fecha_fin_filtro', '')

    # Query inicial
    materiales = Material.objects.filter(activo=True)


    # Excluir materiales dañados
    materiales = materiales.exclude(
        itemsolicitud__estado_ingreso='DAN',
        itemsolicitud__fecha_devolucion_real__isnull=False
    )

    # Aplicar filtros
    if q:
        materiales = materiales.filter(
            Q(nom_material__icontains=q) |
            Q(marca__nom_marca__icontains=q) |
            Q(categoria__nombre_categoria__icontains=q) |
            Q(tipo_material__nombre_tipo_material__icontains=q)
        )
    if marca:
        materiales = materiales.filter(marca__nom_marca=marca)
    if tipo:
        materiales = materiales.filter(tipo_material__nombre_tipo_material=tipo)
    if categoria:
        materiales = materiales.filter(categoria__nombre_categoria=categoria)
        
    if fecha_inicio_str and fecha_fin_str:
        try:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
            fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()

            margen_dias = timedelta(days=2)
            fecha_disponible_limite = fecha_inicio + margen_dias

            materiales_reservados_ids = ItemSolicitud.objects.filter(
                Q(solicitud__estado__in=['PEND', 'APR']) |
                Q(solicitud__estado='PAR', aprobado=True)
            ).filter(
                solicitud__fecha_retiro__lte=fecha_fin,
                solicitud__fecha_devolucion__gte=fecha_inicio
            ).filter(
                Q(
                    # No ha sido devuelto y se espera que esté ocupado después del margen
                    fecha_devolucion_real__isnull=True,
                    solicitud__fecha_devolucion__gt=fecha_disponible_limite
                ) |
                Q(
                    # Ya fue devuelto pero después del margen
                    fecha_devolucion_real__gt=fecha_disponible_limite
                )
            ).values_list('material__id_material', flat=True).distinct()

            materiales = materiales.exclude(id_material__in=materiales_reservados_ids)

        except ValueError:
            pass


    # El resto igual
    marcas = Material.objects.values_list('marca__nom_marca', flat=True).distinct()
    tipos = Material.objects.values_list('tipo_material__nombre_tipo_material', flat=True).distinct()
    categorias_stock = CategoriaDj.objects.all().order_by('nombre_categoria')

    carrito, _ = Carrito.objects.get_or_create(usuario=request.user)
    materiales_en_carrito = set(item.material.id_material for item in carrito.items.all())

    reserva_info = {}

    if fecha_inicio_str and fecha_fin_str:
        # Hay filtro por fechas: usamos el rango
        try:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
            fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
        except ValueError:
            fecha_inicio = None
            fecha_fin = None

        if fecha_inicio and fecha_fin:
            margen_dias = timedelta(days=0)
            fecha_inicio_margen = fecha_inicio - margen_dias

            # Obtenemos reservas activas o futuras en el rango extendido con margen
            reservas_rango = ItemSolicitud.objects.filter(
                Q(solicitud__estado__in=['PEND', 'APR']) |
                Q(solicitud__estado='PAR', aprobado=True),
                solicitud__fecha_retiro__lte=fecha_fin,
                solicitud__fecha_devolucion__gte=fecha_inicio_margen,
                fecha_devolucion_real__isnull=True,
            )

            for item in reservas_rango:
                mat_id = item.material.id_material
                estado_solicitud = item.solicitud.estado
                fecha_dev = item.solicitud.fecha_devolucion
                disponible_desde = fecha_dev + timedelta(days=1)

                # Si la devolución está dentro del margen antes del filtro, o dentro del filtro
                if fecha_dev >= fecha_inicio_margen:
                    if estado_solicitud == 'PEND':
                        reserva_info[mat_id] = "Reserva activa"
                    else:
                        # APR o PAR aprobados
                        # Mostrar disponible desde solo si es posterior al filtro inicio para que sea relevante
                        if disponible_desde >= fecha_inicio:
                            # Si ya hay info y es fecha más lejana, actualizar solo si es menor
                            if mat_id in reserva_info:
                                try:
                                    fecha_str = reserva_info[mat_id].split()[-1]
                                    fecha_existente = datetime.strptime(fecha_str, '%d-%m-%Y').date()
                                    if disponible_desde < fecha_existente:
                                        reserva_info[mat_id] = f"Disponible desde {disponible_desde.strftime('%d-%m-%Y')}"
                                except Exception:
                                    reserva_info[mat_id] = f"Disponible desde {disponible_desde.strftime('%d-%m-%Y')}"
                            else:
                                reserva_info[mat_id] = f"Disponible desde {disponible_desde.strftime('%d-%m-%Y')}"
                else:
                    # Si devolución es anterior al margen, no mostrar nada porque ya está disponible

                    pass

    else:
        # No hay filtro por fecha: usamos lógica actual solo para hoy
        reservas_activas = ItemSolicitud.objects.filter(
            solicitud__fecha_retiro__lte=hoy,
            fecha_devolucion_real__isnull=True,
            solicitud__fecha_devolucion__gte=hoy,
            solicitud__estado__in=['PEND', 'APR', 'PAR']
        )

        for item in reservas_activas:
            mat_id = item.material.id_material
            estado_solicitud = item.solicitud.estado
            fecha_dev = item.solicitud.fecha_devolucion
            disponible_desde = fecha_dev + timedelta(days=1)

            if estado_solicitud == 'PAR':
                if getattr(item, 'aprobado', False):
                    fecha_existente = None
                    if mat_id in reserva_info:
                        try:
                            fecha_str = reserva_info[mat_id].split()[-1]
                            fecha_existente = datetime.strptime(fecha_str, '%d-%m-%Y').date()
                        except:
                            fecha_existente = None

                    if not fecha_existente or disponible_desde > fecha_existente:
                        reserva_info[mat_id] = f"Disponible desde {disponible_desde.strftime('%d-%m-%Y')}"

            elif estado_solicitud == 'APR':
                fecha_existente = None
                if mat_id in reserva_info:
                    try:
                        fecha_str = reserva_info[mat_id].split()[-1]
                        fecha_existente = datetime.strptime(fecha_str, '%d-%m-%Y').date()
                    except:
                        fecha_existente = None

                if not fecha_existente or disponible_desde > fecha_existente:
                    reserva_info[mat_id] = f"Disponible desde {disponible_desde.strftime('%d-%m-%Y')}"

            elif estado_solicitud == 'PEND':
                if mat_id not in reserva_info:
                    reserva_info[mat_id] = "Reserva activa"


    return render(request, 'paginas/crud_material/listar_materiales.html', {
        'materiales': materiales,
        'materiales_en_carrito': materiales_en_carrito,
        'reserva_info': reserva_info,
        'marcas': marcas,
        'tipos': tipos,
        'categorias_stock': categorias_stock,
        'fecha_inicio': fecha_inicio_str,
        'fecha_fin': fecha_fin_str,
        'q': q,
        'marca': marca,
        'tipo': tipo,
        'categoria': categoria,
    })



@login_required
@user_passes_test(is_admin)
def listar_materiales_danados(request):
    # Traer todos los items con estado_ingreso 'DAN'
    items_danados = ItemSolicitud.objects.filter(estado_ingreso='DAN', fecha_devolucion_real__isnull=False)

    return render(request, 'paginas/panel_admin/materiales_danados.html', {
        'items_danados': items_danados
    })


@login_required
@user_passes_test(is_admin)
def reparar_material(request, item_id):
    item = get_object_or_404(ItemSolicitud, id=item_id, estado_ingreso='DAN')
    
    # Cambiar estado_ingreso a 'UTI' (o 'SIN' si prefieres)
    item.estado_ingreso = 'UTI'
    item.save()



    messages.success(request, f"Material '{item.material.nom_material}' devuelto a inventario correctamente.")
    return redirect('listar_materiales_danados')

from django.http import JsonResponse
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.urls import reverse



@login_required
def agregar_al_carrito(request, material_id):
    if request.method == 'POST':
        material = get_object_or_404(Material, pk=material_id)

#       hoy = date.today()
#
#        esta_solicitado = ItemSolicitud.objects.filter(
#            material=material,
#            solicitud__estado='APR',
#            solicitud__fecha_retiro__lte=hoy,
#            solicitud__fecha_devolucion__gte=hoy,
#            fecha_devolucion_real__isnull=True
#        ).exists()
#
#        if esta_solicitado:
#            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
#                return JsonResponse({'estado': 'error', 'mensaje': 'Este material ya ha sido solicitado por otra persona y aún no ha sido devuelto.'})
#           messages.warning(request, "Este material ya ha sido solicitado por otra persona y aún no ha sido devuelto.")
#            return redirect('listar_materiales')

        carrito, creado = Carrito.objects.get_or_create(usuario=request.user)
        item, creado = ItemCarrito.objects.get_or_create(carrito=carrito, material=material)

        if not creado:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'estado': 'error', 'mensaje': 'Este material ya está en tu carrito.'})
          
        else:
            item.cantidad = 1
            item.save()

        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({
                'estado': 'agregado',
                'url_quitar': reverse('quitar_del_carrito', args=[material_id])
            })

        return redirect('listar_materiales')


@login_required
def aprobar_solicitud(request, solicitud_id):
    solicitud = get_object_or_404(Solicitud, id=solicitud_id)
    solicitud.estado = 'APR'
    solicitud.save()

    # Actualizar cada material con la fecha de devolución
    for item in solicitud.itemsolicitud_set.all():
        material = item.material
        if solicitud.fecha_devolucion:
            material.reservado_hasta = solicitud.fecha_devolucion
            material.save()




def ver_carrito(request):
    carrito, creado = Carrito.objects.get_or_create(usuario=request.user)
    return render(request, 'paginas/carrito/ver_carrito.html', {'carrito': carrito})



def vaciar_carrito(request):
    try:
        carrito = Carrito.objects.get(usuario=request.user)
        ItemCarrito.objects.filter(carrito=carrito).delete()
    except Carrito.DoesNotExist:
        pass  # No hay carrito aún

    return redirect('ver_carrito')  # O a donde quieras redirigir




from django.contrib import messages
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, redirect

from django.urls import reverse

@login_required
def quitar_del_carrito(request, material_id):
    if request.method == 'POST':
        carrito = get_object_or_404(Carrito, usuario=request.user)
        try:
            item = ItemCarrito.objects.get(carrito=carrito, material_id=material_id)
            item.delete()
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({
                    'estado': 'quitado',
                    'url_agregar': reverse('agregar_al_carrito', args=[material_id])
                })
            messages.success(request, "Material eliminado del carrito.")
            return redirect('listar_materiales')
        except ItemCarrito.DoesNotExist:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'estado': 'error', 'mensaje': 'Material no encontrado en el carrito.'})
            messages.error(request, "Material no encontrado en el carrito.")
            return redirect('listar_materiales')
    else:
        return JsonResponse({'estado': 'error', 'mensaje': 'Método no permitido'}, status=405)



@login_required 
def crear_solicitud(request):
    hoy = now().date()
    carrito, _ = Carrito.objects.get_or_create(usuario=request.user)

    # Obtener fechas desde sesión, si existen
    fecha_inicio_sesion = request.session.get('fecha_inicio_filtro')
    fecha_fin_sesion = request.session.get('fecha_fin_filtro')

    if request.method == 'POST':
        form = SolicitudForm(request.POST)
        if form.is_valid():
            fecha_retiro = form.cleaned_data['fecha_retiro']
            fecha_devolucion = form.cleaned_data['fecha_devolucion']

            if not fecha_retiro or not fecha_devolucion:
                messages.error(request, "Debes completar las fechas de retiro y devolución.")
                return redirect('crear_solicitud')

            # Validar solapamientos para cada material del carrito
            materiales_solapados = []
            for item_carrito in carrito.items.all():
                existe_solapamiento = ItemSolicitud.objects.filter(
                    material=item_carrito.material,
                    solicitud__estado__in=['PEND', 'APR', 'PAR'],
                    fecha_devolucion_real__isnull=True
                ).filter(
                    solicitud__fecha_retiro__lte=fecha_devolucion,
                    solicitud__fecha_devolucion__gte=fecha_retiro,
                ).exists()

                if existe_solapamiento:
                    materiales_solapados.append(item_carrito.material.nom_material)

            if materiales_solapados:
                nombres = ", ".join(materiales_solapados)
                messages.error(request, f"No se puede crear la solicitud porque estos materiales están reservados en las fechas indicadas: {nombres}")
                return redirect('index')

            # Guardar fechas en sesión para que persistan
            request.session['fecha_inicio_filtro'] = fecha_retiro.isoformat()
            request.session['fecha_fin_filtro'] = fecha_devolucion.isoformat()

            # Crear la solicitud
            solicitud = form.save(commit=False)
            solicitud.usuario = request.user
            solicitud.save()

            for item_carrito in carrito.items.all():
                ItemSolicitud.objects.create(
                    solicitud=solicitud,
                    material=item_carrito.material,
                    cantidad=item_carrito.cantidad
                )
            carrito.items.all().delete()

            enviar_correo_admin(solicitud)

            messages.success(request, "Solicitud enviada correctamente.")
            return redirect('index')

    else:
        initial_data = {}
        if fecha_inicio_sesion:
            initial_data['fecha_retiro'] = fecha_inicio_sesion
        if fecha_fin_sesion:
            initial_data['fecha_devolucion'] = fecha_fin_sesion

        form = SolicitudForm(initial=initial_data)

    return render(request, 'paginas/solicitudes/crear_solicitud.html', {
        'form': form,
        'hoy': hoy,
        'carrito': carrito,
        'items_carrito': carrito.items.all()
    })



def listar_solicitudes(request):
    solicitudes = Solicitud.objects.filter(usuario=request.user).order_by('-fecha_solicitud')
    return render(request, 'paginas/solicitudes/listar_solicitudes.html', {'solicitudes': solicitudes})    



def is_admin(user):
    return user.is_superuser

@login_required
@user_passes_test(is_admin)
def gestionar_solicitud(request, solicitud_id):
    solicitud = get_object_or_404(Solicitud, id=solicitud_id)

    if not (request.user.is_staff or solicitud.usuario == request.user):
        messages.error(request, 'No tienes permiso para gestionar esta solicitud.')
        return redirect('detalle_solicitud', id=solicitud.id)

    if request.method == 'POST':
        form = GestionarSolicitudForm(request.POST, instance=solicitud)
        items_aprobados = request.POST.getlist('aprobados')  # IDs de ítems aprobados

        if form.is_valid():
            solicitud_actualizada = form.save(commit=False)

            for item in solicitud.items.all():
                if str(item.id) in items_aprobados:
                    categoria = item.material.categoria
                    if categoria.stock < item.cantidad:
                        messages.error(
                            request,
                            f"No hay stock suficiente para {item.material.nom_material}. "
                            f"Stock disponible: {categoria.stock}, requerido: {item.cantidad}"
                        )
                        return redirect('gestionar_solicitud', solicitud_id=solicitud.id)
                    item.aprobado = True
                else:
                    item.aprobado = False

                item.rechazado = not item.aprobado
                item.save()

            solicitud_actualizada.comentario_respuesta = form.cleaned_data.get('comentario_respuesta')
            solicitud_actualizada.save()
            solicitud.actualizar_estado()

            enviar_correo_usuario_respuesta(solicitud)

            # Mensaje según el resultado
            total_items = solicitud.items.count()
            total_aprobados = sum(1 for item in solicitud.items.all() if item.aprobado)

            if total_aprobados == total_items:
                messages.success(request, '✅ La solicitud fue aprobada completamente.')
            elif total_aprobados == 0:
                messages.error(request, '❌ La solicitud fue rechazada.')
            else:
                messages.warning(request, '⚠️ La solicitud fue aprobada parcialmente.')

            return redirect('control_solicitudes')
    else:
        form = GestionarSolicitudForm(instance=solicitud)

    return render(request, 'paginas/solicitudes/gestionar_solicitud.html', {
        'form': form,
        'solicitud': solicitud,
        'items': solicitud.items.all(),
    })


def detalle_solicitud(request, numero_solicitud):
    solicitud = get_object_or_404(Solicitud, numero_solicitud=numero_solicitud)
    return render(request, 'paginas/solicitudes/detalle_solicitud.html', {'solicitud': solicitud})

@login_required
def listar_solicitudes(request):
    if request.user.is_staff:
        solicitudes = Solicitud.objects.all().order_by('-fecha_solicitud')
    else:
        solicitudes = Solicitud.objects.filter(usuario=request.user).order_by('-fecha_solicitud')

    return render(request, 'paginas/solicitudes/listar_solicitudes.html', {
        'solicitudes': solicitudes
    })
    
 
@login_required

@login_required
def control_admin_solicitud(request):
    if not request.user.is_staff:
        messages.error(request, "No tienes permiso para acceder a esta página.")
        return redirect('listar_solicitudes')

    estado = request.GET.get('estado', 'PEND').upper()
    estados_validos = ['PEND', 'APR', 'PAR', 'RECH', 'FIN', 'CAN']

    solicitudes_queryset = Solicitud.objects.all()

    # Filtros adicionales
    usuario_filtro = request.GET.get('usuario', '')
    numero_filtro = request.GET.get('numero_solicitud', '')
    fecha_filtro = request.GET.get('fecha_solicitud', '')

    if estado in estados_validos:
        solicitudes_queryset = solicitudes_queryset.filter(estado=estado)

    if usuario_filtro:
        solicitudes_queryset = solicitudes_queryset.filter(usuario__username__icontains=usuario_filtro)
    if numero_filtro:
        solicitudes_queryset = solicitudes_queryset.filter(numero_solicitud__icontains=numero_filtro)
    if fecha_filtro:
        solicitudes_queryset = solicitudes_queryset.filter(fecha_solicitud__date=fecha_filtro)

    solicitudes_queryset = solicitudes_queryset.order_by('-fecha_solicitud')
    paginator = Paginator(solicitudes_queryset, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Lista de estados para usar en el template
    estados = [
        ('PEND', 'Pendientes'),
        ('APR', 'Aprobadas'),
        ('PAR', 'Aprobadas Parcialmente'),
        ('RECH', 'Rechazadas'),
        ('FIN', 'Finalizadas'),
        ('CAN', 'Canceladas'),
    ]

    return render(request, 'paginas/solicitudes/controladminsolicitud.html', {
        'page_obj': page_obj,
        'estado_filtrado': estado,
        'usuario_filtro': usuario_filtro,
        'numero_filtro': numero_filtro,
        'fecha_filtro': fecha_filtro,
        'estados': estados
    })


@login_required
@user_passes_test(is_admin)
def gestionar_devolucion(request, item_id):
    item = get_object_or_404(ItemSolicitud, id=item_id)
    if request.method == 'POST':
        form = DevolverItemForm(request.POST, instance=item)
        if form.is_valid():
            item_devuelto = form.save(commit=False)

            # Establecer fecha de devolución real como la fecha actual
            item_devuelto.fecha_devolucion_real = now().date()
            item_devuelto.save()

            # Verificar si todos los ítems están devueltos
            solicitud = item_devuelto.solicitud
            hay_no_devuelto = solicitud.items.filter(aprobado=True, fecha_devolucion_real__isnull=True).exists()



            if not hay_no_devuelto:
                marcar_solicitud_finalizada(solicitud)

            messages.success(request, "Devolución registrada correctamente.")
            return redirect('gestionar_devoluciones')
    else:
        form = DevolverItemForm(instance=item)

    return render(request, 'paginas/devoluciones/gestionar_devolucion.html', {
        'form': form,
        'item': item
    })
    
    
@login_required
@user_passes_test(is_admin)
def gestionar_devoluciones(request):
    solicitudes = Solicitud.objects.filter(estado__in=['APR', 'PAR']).prefetch_related('items__material')

    solicitudes_con_pendientes = []
    for solicitud in solicitudes:
        # Filtrar ítems aprobados y no devueltos
        items_pendientes = solicitud.items.filter(aprobado=True, fecha_devolucion_real__isnull=True)
        if items_pendientes.exists():
            solicitud.items_pendientes = items_pendientes
            solicitudes_con_pendientes.append(solicitud)

    solicitudes_ordenadas = sorted(
        solicitudes_con_pendientes,
        key=lambda s: s.fecha_devolucion or s.fecha_retiro
    )

    return render(request, 'paginas/devoluciones/gestionar_devoluciones.html', {
        'solicitudes': solicitudes_ordenadas
    })




@login_required
@user_passes_test(is_admin)
def listar_usuarios(request):
    usuarios = User.objects.all()
    return render(request, 'paginas/admin/listar_usuarios.html', {'usuarios': usuarios})




@login_required
@user_passes_test(is_admin)
def desactivar_usuario(request, usuario_id):
    usuario = get_object_or_404(User, id=usuario_id)
    if usuario != request.user:  # No dejar que se desactive a sí mismo
        usuario.is_active = False
        usuario.save()
    return redirect('listar_usuarios')

@login_required
@user_passes_test(is_admin)
def activar_usuario(request, usuario_id):
    usuario = get_object_or_404(User, id=usuario_id)
    usuario.is_active = True
    usuario.save()
    return redirect('listar_usuarios')




@login_required
@user_passes_test(is_admin)
def ver_solicitudes_usuario(request, user_id):
    usuario = get_object_or_404(User, id=user_id)
    solicitudes = Solicitud.objects.filter(usuario=usuario).order_by('-fecha_solicitud')
    return render(request, 'paginas/admin/solicitudes_usuario.html', {
        'usuario': usuario,
        'solicitudes': solicitudes
    })

@login_required
@user_passes_test(is_admin)
def reporte_prestamos(request):
    solicitudes = Solicitud.objects.filter(estado__in=['APR', 'FIN', 'PAR']).select_related('usuario').prefetch_related('items__material').order_by('-fecha_solicitud')

    fecha_inicio = request.GET.get('inicio')
    fecha_fin = request.GET.get('fin')
    usuario_id = request.GET.get('usuario')
    hoy = date.today().isoformat()
    if fecha_inicio:
        solicitudes = solicitudes.filter(fecha_solicitud__date__gte=fecha_inicio)
    if fecha_fin:
        solicitudes = solicitudes.filter(fecha_solicitud__date__lte=fecha_fin)
    if usuario_id:
        solicitudes = solicitudes.filter(usuario__id=usuario_id)

    # Filtramos los ítems para mostrar en el reporte
    for solicitud in solicitudes:
        if solicitud.estado in ['PAR', 'APR', 'FIN']:
            solicitud.items_filtrados = solicitud.items.filter(aprobado=True)
        else:
            solicitud.items_filtrados = solicitud.items.all()  # o .all() si quieres mostrar todo en otros estados


    usuarios = get_user_model().objects.all()

    context = {
        'solicitudes': solicitudes,
        'usuarios': usuarios,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'usuario_id': usuario_id,
        'hoy': hoy,
    }
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        
        return render(request, 'paginas/reportes/_lista_solicitudes.html', context)
    else:
        return render(request, 'paginas/reportes/reporte_prestamos.html', context)

    
from django.utils.dateparse import parse_date

@login_required
@user_passes_test(is_admin)
def exportar_reporte_pdf(request):
    solicitudes = Solicitud.objects.filter(estado__in=['APR', 'FIN', 'PAR']).select_related('usuario').prefetch_related('items__material').order_by('-fecha_solicitud')

    fecha_inicio = request.GET.get('inicio')
    fecha_fin = request.GET.get('fin')
    usuario_id = request.GET.get('usuario')

    if fecha_inicio:
        fecha_inicio = parse_date(fecha_inicio)
        if fecha_inicio:
            solicitudes = solicitudes.filter(fecha_solicitud__date__gte=fecha_inicio)

    if fecha_fin:
        fecha_fin = parse_date(fecha_fin)
        if fecha_fin:
            solicitudes = solicitudes.filter(fecha_solicitud__date__lte=fecha_fin)

    if usuario_id and usuario_id.isdigit():
        solicitudes = solicitudes.filter(usuario__id=int(usuario_id))


    for solicitud in solicitudes:
        if solicitud.estado in ['PAR', 'APR', 'FIN']:
            solicitud.items_filtrados = solicitud.items.filter(aprobado=True)
        else:
            solicitud.items_filtrados = solicitud.items.none()

    template = get_template('paginas/reportes/reporte_prestamos_pdf.html')
    html = template.render({'solicitudes': solicitudes})

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_prestamos.pdf"'

    pisa.CreatePDF(BytesIO(html.encode('UTF-8')), dest=response, encoding='UTF-8')
    return response


def marcar_solicitud_finalizada(solicitud):
    if solicitud.estado != 'FIN':
        solicitud.estado_anterior = solicitud.estado
        solicitud.estado = 'FIN'
        solicitud.save()


@login_required
def editar_perfil(request):
    if request.method == 'POST':
        perfil_form = EditarPerfilForm(request.POST, request.FILES, instance=request.user)
        password_form = PasswordChangeForm(user=request.user, data=request.POST)

        if perfil_form.is_valid():
            perfil_form.save()
            # Si el usuario ingresó datos en el formulario de contraseña:
            if password_form.is_valid() and password_form.cleaned_data.get('new_password1'):
                password_form.save()
                messages.success(request, '¡Contraseña actualizada! Por favor, inicia sesión de nuevo.')
                return redirect('login')  # Redirigir al login si cambió la contraseña

            update_session_auth_hash(request, request.user)  # Mantiene la sesión activa
            messages.success(request, '¡Perfil actualizado correctamente!')
            return redirect('perfil')  # Aquí rediriges a la página de perfil

        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')

    else:
        perfil_form = EditarPerfilForm(instance=request.user)
        password_form = PasswordChangeForm(user=request.user)

    return render(request, 'paginas/inicio/editar_perfil.html', {
        'perfil_form': perfil_form,
        'password_form': password_form,
    })
    





@login_required
def cancelar_solicitud(request, solicitud_id):
    solicitud = get_object_or_404(Solicitud, id=solicitud_id, usuario=request.user)

    if solicitud.estado == 'PEND':
        solicitud.estado = 'CAN'
        solicitud.save()

        # Actualiza los ítems a rechazado
        solicitud.items.all().update(aprobado=False, rechazado=True)

        messages.success(request, 'Solicitud cancelada exitosamente.')
    else:
        messages.error(request, 'Solo puedes cancelar solicitudes pendientes.')

    return redirect('listar_solicitudes')  # o la vista que estés usando para listarlas





from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import HttpResponse
from datetime import date
import openpyxl
from openpyxl.styles import Font

@login_required
@user_passes_test(is_admin)
def reporte_excel(request):
    solicitudes = Solicitud.objects.filter(estado__in=['APR', 'FIN', 'PAR', 'PEND', 'RECH']) \
        .select_related('usuario') \
        .prefetch_related('items__material') \
        .order_by('-fecha_solicitud')

    fecha_inicio = request.GET.get('inicio')
    fecha_fin = request.GET.get('fin')
    usuario_id = request.GET.get('usuario')

    # Validar y limpiar valores 'None' o vacíos
    if fecha_inicio in [None, '', 'None']:
        fecha_inicio = None
    if fecha_fin in [None, '', 'None']:
        fecha_fin = None
    if usuario_id in [None, '', 'None']:
        usuario_id = None

    if fecha_inicio:
        solicitudes = solicitudes.filter(fecha_solicitud__date__gte=fecha_inicio)
    if fecha_fin:
        solicitudes = solicitudes.filter(fecha_solicitud__date__lte=fecha_fin)
    if usuario_id:
        solicitudes = solicitudes.filter(usuario__id=usuario_id)

    import openpyxl
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Préstamos"

    headers = ['ID Solicitud', 'Fecha Solicitud', 'Estado', 'Usuario', 'Fecha Retiro', 'Fecha Devolución',
               'Material', 'Cantidad', 'Estado Solicitud', 'Fecha Devolución Real', 'Estado Ingreso', 'Ubicación']
    ws.append(headers)

    for solicitud in solicitudes:
        # Mostrar todos los items sin filtrar por aprobado
        for item in solicitud.items.all():
            estado_item = 'Aprobado' if item.aprobado else 'Pendiente'
            if item.rechazado:
                estado_item = 'Rechazado'

            # Ajustar valores para items rechazados
            fecha_devolucion_real = 'No aplica' if item.rechazado else (item.fecha_devolucion_real.strftime('%d-%m-%Y') if item.fecha_devolucion_real else 'Pendiente')
            estado_ingreso = 'No aplica' if item.rechazado else (item.get_estado_ingreso_display() or 'Pendiente')

            row = [
                solicitud.numero_solicitud,
                solicitud.fecha_solicitud.strftime('%d-%m-%Y %H:%M'),
                solicitud.get_estado_display(),
                solicitud.usuario.get_full_name() or solicitud.usuario.username,
                solicitud.fecha_retiro.strftime('%d-%m-%Y') if solicitud.fecha_retiro else '',
                solicitud.fecha_devolucion.strftime('%d-%m-%Y') if solicitud.fecha_devolucion else '',
                item.material.nom_material if item.material else '',
                item.cantidad,
                estado_item,
                fecha_devolucion_real,
                estado_ingreso,
                solicitud.ubicacion_solicitud or '',
            ]
            ws.append(row)

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = max_length + 2
        ws.column_dimensions[col_letter].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=reporte_prestamos.xlsx'
    wb.save(response)
    return response



from django.views.decorators.http import require_POST
from django.http import JsonResponse
from datetime import datetime

@require_POST
def verificar_reservas_ajax(request):
    fecha_retiro = request.POST.get('fecha_retiro')
    fecha_devolucion = request.POST.get('fecha_devolucion')

    if not fecha_retiro or not fecha_devolucion:
        return JsonResponse({'error': 'Fechas no proporcionadas.'}, status=400)

    try:
        fecha_retiro = datetime.strptime(fecha_retiro, '%Y-%m-%d').date()
        fecha_devolucion = datetime.strptime(fecha_devolucion, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({'error': 'Formato de fecha inválido.'}, status=400)

    carrito, _ = Carrito.objects.get_or_create(usuario=request.user)
    materiales_solapados = []

    for item_carrito in carrito.items.all():
        existe_solapamiento = ItemSolicitud.objects.filter(
            material=item_carrito.material,
            solicitud__estado__in=['PEND', 'APR', 'PAR'],
            fecha_devolucion_real__isnull=True
        ).filter(
            solicitud__fecha_retiro__lte=fecha_devolucion,
            solicitud__fecha_devolucion__gte=fecha_retiro,
        ).exists()

        if existe_solapamiento:
            materiales_solapados.append(item_carrito.material.nom_material)

    return JsonResponse({'conflictos': materiales_solapados})


from collections import defaultdict
from datetime import datetime
from django.contrib import messages
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required

def calcular_disponibilidad(material, fecha_inicio, fecha_fin):
    return not material.itemsolicitud_set.filter(
        solicitud__estado__in=['APR', 'PAR', 'PEND'],
        solicitud__fecha_retiro__lte=fecha_fin,
        solicitud__fecha_devolucion__gte=fecha_inicio,
        rechazado=False
    ).exists()

from datetime import datetime, date

@login_required
def seleccion_masiva_materiales(request):
    categoria_id = request.GET.get('categoria')
    marca_id = request.GET.get('marca')
    tipo_id = request.GET.get('tipo')


    # Inicializar fechas
    fecha_retiro = None
    fecha_devolucion = None

    # Manejar fecha_retiro
    if 'fecha_retiro' in request.GET:
        valor = request.GET.get('fecha_retiro')
        if valor:
            fecha_retiro = valor
            request.session['fecha_inicio_filtro'] = valor
        else:
            request.session.pop('fecha_inicio_filtro', None)  # Borrar de sesión si vacío

    # Manejar fecha_devolucion
    if 'fecha_devolucion' in request.GET:
        valor = request.GET.get('fecha_devolucion')
        if valor:
            fecha_devolucion = valor
            request.session['fecha_fin_filtro'] = valor
        else:
            request.session.pop('fecha_fin_filtro', None)

    # Si no vienen en el GET, usar lo guardado en la sesión
    if not fecha_retiro:
        fecha_retiro = request.session.get('fecha_inicio_filtro')
    if not fecha_devolucion:
        fecha_devolucion = request.session.get('fecha_fin_filtro')

    # Si aún no hay fechas, usar hoy como predeterminado
    if not fecha_retiro or not fecha_devolucion:
        hoy = date.today()
        fecha_retiro = hoy.isoformat()
        fecha_devolucion = hoy.isoformat()



    # Si aún no hay fechas, usar hoy como predeterminado
    if not fecha_retiro or not fecha_devolucion:
        hoy = date.today()
        fecha_retiro = hoy.isoformat()
        fecha_devolucion = hoy.isoformat()

    try:
        retiro = datetime.strptime(fecha_retiro, "%Y-%m-%d").date()
        devolucion = datetime.strptime(fecha_devolucion, "%Y-%m-%d").date()
    except ValueError:
        retiro = devolucion = date.today()

    # Validar que la fecha de devolución no sea anterior a la de retiro
    if devolucion < retiro:
        messages.error(request, "La fecha de devolución no puede ser anterior a la fecha de retiro.")


    materiales = Material.objects.all()

    if categoria_id:
        materiales = materiales.filter(categoria__id_categoria=categoria_id)
    if marca_id:
        materiales = materiales.filter(marca__id_marca=marca_id)
    if tipo_id:
        materiales = materiales.filter(tipo_material__id_tipo_material=tipo_id)

    # Filtrar solo materiales disponibles en rango
    materiales_filtrados = [
        m for m in materiales if calcular_disponibilidad(m, retiro, devolucion)
    ]

    # Agrupar materiales
    grupos = defaultdict(list)
    for mat in materiales_filtrados:
        key = (mat.categoria.id_categoria, mat.marca.id_marca, mat.tipo_material.id_tipo_material, mat.nom_material)
        grupos[key].append(mat)

    grupos_agrupados = []
    for key, mats in grupos.items():
        grupos_agrupados.append({
            'categoria': mats[0].categoria,
            'marca': mats[0].marca,
            'tipo': mats[0].tipo_material,
            'nombre': mats[0].nom_material,
            'cantidad': len(mats),
            'materiales': mats,
        })

    # Procesar POST
    if request.method == 'POST':
        total_solicitado = 0
        for grupo in grupos_agrupados:
            grupo_id = grupo['materiales'][0].id_material
            cantidad_str = request.POST.get(f'cantidad_{grupo_id}')
            try:
                cantidad = int(cantidad_str)
            except (TypeError, ValueError):
                cantidad = 0

            if cantidad > grupo['cantidad']:
                messages.error(request, f"No puedes solicitar más que {grupo['cantidad']} del grupo {grupo['nombre']}.")
                return redirect(request.path + f"?categoria={categoria_id}&marca={marca_id}&tipo={tipo_id}&fecha_retiro={fecha_retiro}&fecha_devolucion={fecha_devolucion}")

            if cantidad > 0:
                for mat in grupo['materiales'][:cantidad]:
                    agregar_al_carrito(request, mat.id_material)
                total_solicitado += cantidad

        if total_solicitado > 0:
            return redirect('ver_carrito')
        else:
            messages.error(request, 'No se seleccionó ninguna cantidad válida.')

    context = {
        'grupos_agrupados': grupos_agrupados,
        'fecha_retiro': fecha_retiro,
        'fecha_devolucion': fecha_devolucion,
        'categorias': CategoriaDj.objects.all(),
        'marcas': Marca.objects.all(),
        'tipos': TipoMaterial.objects.all(),
        'categoria_id': categoria_id or '',
        'marca_id': marca_id or '',
        'tipo_id': tipo_id or '',
    }
    return render(request, 'paginas/inventario/seleccion_masiva.html', context)


def detalle_material(request, id):
    material = get_object_or_404(Material, id_material=id)
    return render(request, 'paginas/detalle/detalle_material.html', {'material': material})


@login_required
def admin_listar_materiales(request):
    materiales = Material.objects.filter(activo=True).order_by('categoria', 'nom_material')

    # Obtener filtros de GET o sesión
    q_nombre = request.GET.get('q_nombre') or request.session.get('q_nombre', '')
    q_marca = request.GET.get('q_marca') or request.session.get('q_marca', '')
    q_modelo = request.GET.get('q_modelo') or request.session.get('q_modelo', '')
    q_serie = request.GET.get('q_serie') or request.session.get('q_serie', '')

    # Limpiar filtros si no hay GET
    if not request.GET:
        request.session.pop('q_nombre', None)
        request.session.pop('q_marca', None)
        request.session.pop('q_modelo', None)
        request.session.pop('q_serie', None)
    else:
        # Guardar filtros en sesión
        request.session['q_nombre'] = q_nombre
        request.session['q_marca'] = q_marca
        request.session['q_modelo'] = q_modelo
        request.session['q_serie'] = q_serie

    # Aplicar filtros
    if q_nombre:
        materiales = materiales.filter(nom_material__istartswith=q_nombre)
    if q_marca:
        materiales = materiales.filter(marca__nom_marca__istartswith=q_marca)
    if q_modelo:
        materiales = materiales.filter(modelo_material__istartswith=q_modelo)
    if q_serie:
        materiales = materiales.filter(codigo_barra__istartswith=q_serie)

    reserva_info = {}

    return render(request, 'paginas/crud_material/listar_materiales_admin.html', {
        'materiales': materiales,
        'reserva_info': reserva_info,
        'q_nombre': q_nombre,
        'q_marca': q_marca,
        'q_modelo': q_modelo,
        'q_serie': q_serie,
    })

    
    
@login_required
def admin_editar_material(request, pk):
    material = get_object_or_404(Material, pk=pk)

    if request.method == 'POST':
        form = MaterialForm(request.POST, request.FILES, instance=material)
        if form.is_valid():
            form.save()
            messages.success(request, 'Material editado correctamente.')
            return redirect('admin_listar_materiales')
    else:
        form = MaterialForm(instance=material)

    return render(request, 'paginas/crud_material/editar_material_admin.html', {'form': form, 'material': material})



from django.db.models import ProtectedError


@login_required
def admin_eliminar_material(request, pk):
    material = get_object_or_404(Material, pk=pk)

    if request.method == 'POST':
        try:
            material.delete()
            messages.success(request, 'Material eliminado correctamente.')
        except ProtectedError:
            material.activo = False
            material.save()
            messages.warning(request, 'El material no se puede eliminar porque tiene transacciones. Ha sido desactivado y ya no aparecerá en el listado.')
        return redirect('admin_listar_materiales')

    return render(request, 'paginas/crud_material/eliminar_material_admin.html', {'material': material})



@login_required
def admin_listar_materiales_inactivos(request):
    q_nombre = request.GET.get('q_nombre', '').strip()
    q_marca = request.GET.get('q_marca', '').strip()
    q_modelo = request.GET.get('q_modelo', '').strip()
    q_serie = request.GET.get('q_serie', '').strip()

    materiales = Material.objects.filter(activo=False)

    if q_nombre:
        materiales = materiales.filter(nom_material__istartswith=q_nombre)
    if q_marca:
        materiales = materiales.filter(marca__nom_marca__istartswith=q_marca)
    if q_modelo:
        materiales = materiales.filter(modelo_material__istartswith=q_modelo)
    if q_serie:
        materiales = materiales.filter(codigo_barra__istartswith=q_serie)

    context = {
        'materiales': materiales,
        'q_nombre': q_nombre,
        'q_marca': q_marca,
        'q_modelo': q_modelo,
        'q_serie': q_serie,
    }
    return render(request, 'paginas/crud_material/materiales_inactivos.html', context)


@login_required
def reactivar_material(request, pk):
    material = get_object_or_404(Material, pk=pk)
    material.activo = True
    material.save()
    messages.success(request, 'Material reactivado correctamente.')
    return redirect('admin_listar_materiales_inactivos')


@login_required
def detalle_material_admin(request, pk):
    material = get_object_or_404(Material, pk=pk)
    return render(request, 'paginas/crud_material/detalle_material_admin.html', {'material': material})


def editar_carrusel(request):
    slides = SlideCarrusel.objects.all()
    return render(request, 'paginas/panel_admin/panel_admin_carrusel.html', {'slides': slides})

def agregar_slide(request):
    if request.method == 'POST':
        form = SlideCarruselForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('editar_carrusel')
    else:
        form = SlideCarruselForm()
    return render(request, 'paginas/panel_admin/form_slide.html', {'form': form, 'modo': 'Agregar'})

def editar_slide(request, pk):
    slide = get_object_or_404(SlideCarrusel, pk=pk)
    if request.method == 'POST':
        form = SlideCarruselForm(request.POST, request.FILES, instance=slide)
        if form.is_valid():
            form.save()
            return redirect('editar_carrusel')
    else:
        form = SlideCarruselForm(instance=slide)
    return render(request, 'paginas/panel_admin/form_slide.html', {'form': form, 'modo': 'Editar'})

def eliminar_slide(request, pk):
    slide = get_object_or_404(SlideCarrusel, pk=pk)
    slide.delete()
    return redirect('editar_carrusel')
